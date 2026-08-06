"""
Builds the FYIND upload-ready export with DYNAMIC attribute columns:
each distinct attribute name found across the batch (e.g. "Size", "Color",
"Angle") gets its own column, named after the attribute -- not the old
generic "Attribute 1".."Attribute 6" slots.

Special case: "Size" always gets a paired "UOM" column right after it,
with the unit of measurement (e.g. mm, ") split out of the value so the
Size column holds just the numbers. This only applies to Size -- other
attributes (Color, Angle, Material, etc.) keep their value as-is in a
single column, matching the sample_upload_ready_file.xlsx you provided.

Because the number and names of attribute columns vary batch to batch, we
build the header row ourselves (matching the real template's fixed column
text/order exactly) rather than reusing the static template file, which
only had room for a fixed 6 generic attribute slots.
"""
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import config

BASE_DIR = Path(__file__).resolve().parent.parent

# Fixed columns before and after the dynamic attribute block, in the exact
# order/text of the real FYIND template (templates_xlsx/upload_template.xlsx).
FIXED_BEFORE_ATTRS = [
    "Listing Code", "Item Code", "Relationship", "Item Description",
    "Manufacturer", "Place Of Origin", "Additional Information", "Images",
    "Featured Image", "Video File Name ", "Speciality Tags", "Supplier",
    "Supplier Status", "Qoh", "Datasheets", "Stocklist", "Short Description",
    "Price", "Clearance", "Sku", "Attribute Set",
]
FIXED_AFTER_ATTRS = ["Brand", "Warranty", "Guarantee", "Return & Exchange", "Status"]
TRAILING_EXTRA = ["Confidence", "Review"]  # appended beyond the real template, per your spec

# Columns we always populate (the rest are left blank -- filled elsewhere/later)
POPULATED_FIXED = {
    "Item Description", "Manufacturer", "Place Of Origin",
    "Additional Information", "Supplier", "Supplier Status", "Price",
    "Attribute Set", "Brand", "Warranty", "Guarantee", "Return & Exchange", "Status",
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _split_size_uom(value: str):
    """
    Splits a comma-separated Size value like '1/2", 3/4", 1"' or
    '19 mm, 20 mm, 25 mm' into (values_without_unit, unit). Only splits if
    every comma-separated part ends in the SAME unit token -- otherwise
    returns the value unchanged with an empty unit (safer than guessing).
    """
    if not value:
        return value, ""
    parts = value.split(",")
    unit_pattern = re.compile(r'([A-Za-z]+|["\'])\s*$')
    units_found = []
    stripped_parts = []
    for p in parts:
        p_stripped = p.strip()
        m = unit_pattern.search(p_stripped)
        if m:
            units_found.append(m.group(1))
            stripped_parts.append(p_stripped[: m.start()])
        else:
            units_found.append(None)
            stripped_parts.append(p_stripped)
    non_none = [u for u in units_found if u is not None]
    if non_none and all(u == non_none[0] for u in units_found):
        return ", ".join(stripped_parts), non_none[0]
    return value, ""


def _ordered_attribute_names(products: list[dict]) -> list[str]:
    """
    Returns the distinct attribute names used across the whole batch, in
    first-seen order, with "Size" forced to the front if present (so its
    paired "UOM" column reliably lands right after it).
    """
    seen = []
    seen_lower = set()
    for p in products:
        for a in p.get("attributes") or []:
            name = (a.get("name") or "").strip()
            if not name:
                continue
            key = name.lower()
            if key not in seen_lower:
                seen_lower.add(key)
                seen.append(name)

    size_names = [n for n in seen if n.lower() == "size"]
    others = [n for n in seen if n.lower() != "size"]
    return size_names + others  # at most one "Size" entry in practice


def build_export(products: list[dict], attribute_library: dict, supplier_name: str) -> str:
    """
    products: list of dicts as returned by classifier.classify_catalog()
    Returns the filepath (relative to BASE_DIR) of the generated .xlsx file.
    """
    attr_names = _ordered_attribute_names(products)
    has_size = bool(attr_names) and attr_names[0].lower() == "size"

    # Build the dynamic middle section: Size (+UOM) then everything else.
    dynamic_headers = []
    if has_size:
        dynamic_headers.append(attr_names[0])  # canonical "Size" casing as Claude gave it
        dynamic_headers.append("UOM")
        rest = attr_names[1:]
    else:
        rest = attr_names
    dynamic_headers.extend(rest)

    headers = FIXED_BEFORE_ATTRS + dynamic_headers + FIXED_AFTER_ATTRS + TRAILING_EXTRA

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    _style_header_row(ws, len(headers))

    col_letter = {name: get_column_letter(i + 1) for i, name in enumerate(headers)}

    row_idx = 2
    for p in products:
        if not p.get("item_description"):
            continue  # item_description is mandatory; skip anything missing it

        pname = p.get("matched_product_name")
        attribute_set_name = ""
        if pname and pname in attribute_library:
            attribute_set_name = attribute_library[pname]["attribute_set_name"]
        elif pname:
            attribute_set_name = pname

        ws[f"{col_letter['Item Description']}{row_idx}"] = p["item_description"]
        ws[f"{col_letter['Manufacturer']}{row_idx}"] = p.get("manufacturer") or ""
        ws[f"{col_letter['Place Of Origin']}{row_idx}"] = p.get("place_of_origin") or ""
        ws[f"{col_letter['Additional Information']}{row_idx}"] = p.get("additional_information") or ""
        ws[f"{col_letter['Supplier']}{row_idx}"] = supplier_name
        ws[f"{col_letter['Supplier Status']}{row_idx}"] = p.get("supplier_status") or "Suppliers"
        ws[f"{col_letter['Price']}{row_idx}"] = p.get("price") or ""
        ws[f"{col_letter['Attribute Set']}{row_idx}"] = attribute_set_name

        # Map this product's attributes by lowercase name for lookup.
        product_attrs = {}
        for a in p.get("attributes") or []:
            name = (a.get("name") or "").strip()
            if name:
                product_attrs[name.lower()] = a.get("value", "")

        if has_size:
            size_val = product_attrs.pop("size", "")
            if size_val:
                value_only, uom = _split_size_uom(size_val)
                ws[f"{col_letter[attr_names[0]]}{row_idx}"] = value_only
                ws[f"{col_letter['UOM']}{row_idx}"] = uom

        for name in rest:
            val = product_attrs.get(name.lower(), "")
            ws[f"{col_letter[name]}{row_idx}"] = val

        ws[f"{col_letter['Brand']}{row_idx}"] = p.get("brand") or ""
        ws[f"{col_letter['Warranty']}{row_idx}"] = p.get("warranty") or ""
        ws[f"{col_letter['Guarantee']}{row_idx}"] = p.get("guarantee") or ""
        ws[f"{col_letter['Return & Exchange']}{row_idx}"] = p.get("return_exchange") or ""
        ws[f"{col_letter['Status']}{row_idx}"] = "Active"

        confidence = p.get("confidence", 0)
        ws[f"{col_letter['Confidence']}{row_idx}"] = f"{confidence}%"
        ws[f"{col_letter['Review']}{row_idx}"] = (
            "Confident" if confidence >= config.CONFIDENCE_THRESHOLD else "Needs review"
        )

        row_idx += 1

    _style_body(ws, len(headers), row_idx - 1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_supplier = "".join(c if c.isalnum() else "_" for c in supplier_name)[:40]
    out_dir = BASE_DIR / "outputs"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"{safe_supplier}_{timestamp}.xlsx"
    wb.save(out_path)
    return str(out_path.relative_to(BASE_DIR))


def _style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A2"


def _style_body(ws, ncols, nrows):
    for row in ws.iter_rows(min_row=2, max_row=max(nrows, 1), max_col=ncols):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
